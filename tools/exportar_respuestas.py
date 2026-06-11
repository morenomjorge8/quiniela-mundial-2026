"""
Exporta las respuestas de una jornada desde "Respuestas Quiniela 2026.xlsx"
(las pestañas "Form Responses N" del Google Sheet) a data/respuestas/jornada_N.csv,
que es lo que consume data/respuestas_loader.py.

Encuentra la pestaña correcta por el número del primer partido de la jornada
(J1 → #1, J2 → #13, ...), sin depender del nombre de la pestaña.

Uso:
    py tools/exportar_respuestas.py 1
"""
import csv
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

EXCEL_RESP = os.path.join(os.path.dirname(__file__), '..', 'Respuestas Quiniela 2026.xlsx')
DEST_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'respuestas')


def _num_partido(header):
    m = re.match(r'\s*#\s*(\d+)\b', str(header))
    return int(m.group(1)) if m else None


def _hoja_de_jornada(wb, jornada):
    objetivo = (jornada - 1) * 12 + 1
    for name in wb.sheetnames:
        ws = wb[name]
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        nums = [n for n in (_num_partido(h) for h in headers) if n is not None]
        if nums and min(nums) == objetivo:
            return ws
    return None


def exportar(jornada):
    wb = openpyxl.load_workbook(EXCEL_RESP, data_only=True)
    ws = _hoja_de_jornada(wb, jornada)
    if ws is None:
        raise SystemExit(f"No encontré la pestaña de la jornada {jornada} en {EXCEL_RESP}")

    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    # columna del nombre (para filtrar filas vacías)
    col_nombre = next(
        (i for i, h in enumerate(headers) if h and str(h).strip().lower().startswith('tu nombre')),
        1,
    )

    os.makedirs(DEST_DIR, exist_ok=True)
    destino = os.path.join(DEST_DIR, f'jornada_{jornada}.csv')
    n = 0
    with open(destino, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow([('' if h is None else h) for h in headers])
        for r in rows[1:]:
            if len(r) <= col_nombre or r[col_nombre] in (None, ''):
                continue
            w.writerow([('' if c is None else c) for c in r])
            n += 1

    print(f"Escrito {destino} ({n} respuestas)")
    return destino


if __name__ == '__main__':
    jornada = int(sys.argv[1]) if len(sys.argv) > 1 else 1
    exportar(jornada)
