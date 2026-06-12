"""
Lee los resultados reales de una jornada desde "Respuestas Quiniela 2026.xlsx"
(fila "Respuesta" de la pestaña "Form Responses N") y los escribe/mezcla en
data/resultados_reales.json.

Solo registra los partidos que ya tienen resultado (celdas no vacías), así se
puede correr en vivo según van terminando los partidos.

Uso:
    py tools/exportar_resultados.py 1
"""
import json
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from data.respuestas_loader import _parsear_resultado, _parsear_entero

EXCEL_RESP = os.path.join(os.path.dirname(__file__), '..', 'Respuestas Quiniela 2026.xlsx')
RESULTADOS_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'resultados_reales.json')

FILA_RESULTADO = 'respuesta'   # nombre (col "Tu nombre") de la fila con los resultados


def _num_partido(header):
    m = re.match(r'\s*#\s*(\d+)\b', str(header))
    return int(m.group(1)) if m else None


def _hoja_de_jornada(wb, jornada):
    objetivo = (jornada - 1) * 12 + 1
    for name in wb.sheetnames:
        ws = wb[name]
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        nums = [n for n in (_num_partido(h) for h in headers) if n is not None]
        if nums and min(nums) == objetivo:
            return ws
    return None


def exportar(jornada):
    wb = openpyxl.load_workbook(EXCEL_RESP, data_only=True)
    ws = _hoja_de_jornada(wb, jornada)
    if ws is None:
        raise SystemExit(f"No encontré la pestaña de la jornada {jornada} en {EXCEL_RESP}")

    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    col_nombre = next(
        (i for i, h in enumerate(headers) if h and str(h).strip().lower().startswith('tu nombre')),
        1,
    )
    cols_partido = {i: _num_partido(h) for i, h in enumerate(headers) if _num_partido(h)}
    col_rojas = next((i for i, h in enumerate(headers) if h and 'roja' in str(h).lower()), None)
    col_penales = next((i for i, h in enumerate(headers) if h and 'penal' in str(h).lower()), None)

    fila = next(
        (r for r in rows[1:]
         if len(r) > col_nombre and str(r[col_nombre] or '').strip().lower() == FILA_RESULTADO),
        None,
    )
    if fila is None:
        raise SystemExit(
            f"No encontré la fila de resultados (columna 'Tu nombre' == '{FILA_RESULTADO}') "
            f"en la jornada {jornada}."
        )

    resultados = {}
    for i, num in cols_partido.items():
        valor = fila[i] if i < len(fila) else None
        if valor is None or str(valor).strip() == '':
            continue
        resultados[str(num)] = _parsear_resultado(str(valor)).value

    def _entero_opcional(idx, campo):
        if idx is None or idx >= len(fila) or fila[idx] is None or str(fila[idx]).strip() == '':
            return None
        return _parsear_entero(fila[idx], campo)

    total_rojas = _entero_opcional(col_rojas, 'rojas')
    total_penales = _entero_opcional(col_penales, 'penales')

    data = {}
    if os.path.exists(RESULTADOS_PATH):
        with open(RESULTADOS_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
    entrada = {'resultados': resultados}
    if total_rojas is not None:
        entrada['total_rojas'] = total_rojas
    if total_penales is not None:
        entrada['total_penales'] = total_penales
    data[str(jornada)] = entrada

    with open(RESULTADOS_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Jornada {jornada}: {len(resultados)} partidos con resultado "
          f"{sorted(int(k) for k in resultados)}")
    print(f"  rojas={total_rojas}  penales={total_penales}")
    print(f"Escrito {RESULTADOS_PATH}")


if __name__ == '__main__':
    jornada = int(sys.argv[1]) if len(sys.argv) > 1 else 1
    exportar(jornada)
