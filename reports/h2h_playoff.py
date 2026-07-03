"""
Genera un PDF (HTML imprimible) por cada CRUCE head-to-head de playoffs.

Cada archivo muestra, lado a lado, las predicciones de los dos rivales de un
cruce, con sus caricaturas, una columna central de RESULTADO REAL y los PUNTOS
que cada quien gana en cada partido (3 exacto / 2 resultado / +1 primer gol),
más los bonos de la ronda (rojas y penales).

- Cruces: del bracket real (tabla final de la temporada regular).
  · J7 = Cuartos (primeros 8 dieciseisavos, #73–80).
  · J8 = Semifinales (segundos 8 dieciseisavos, #81–88); los cruces salen de
    resolver los cuartos con los puntos de la J7.
- Predicciones: "Respuestas Quiniela 2026 Playoffs.xlsx" (hoja por ronda).
- Resultados reales: fila 'Respuesta' en la hoja de esa ronda (+ override en
  data/resultados_playoffs.json).

Uso:
    py reports/h2h_playoff.py            # genera J7 y J8 (reports/output/)
"""
import base64
import json
import os
import re
import sys
import webbrowser

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl

from data.loader import cargar_participantes
from data.historial_io import cargar_historial_resultados
from quiniela.standings import calcular_tabla_general
from quiniela import playoffs
from quiniela.models import PrediccionPlayoff, ResultadoPlayoff
from quiniela.playoff_scorer import puntos_partido, BONUS_ROJAS, BONUS_PENALES
from reports.generar_reporte import _CARICATURA_FILE

ROOT = os.path.join(os.path.dirname(__file__), '..')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'output')
DOCS_DIR = os.path.join(ROOT, 'docs')
PLAYOFF_XLSX = os.path.join(ROOT, 'Respuestas Quiniela 2026 Playoffs.xlsx')
RESULTADOS_JSON = os.path.join(ROOT, 'data', 'resultados_playoffs.json')
CARICATURAS_DIR = os.path.join(ROOT, 'Caricaturas')

# Partidos por ronda (deben coincidir con forms/crear_forms_playoffs.js).
MATCHES_J7 = [
    (73, 'Sudáfrica', 'Canadá'),
    (74, 'Brasil', 'Japón'),
    (75, 'Alemania', 'Paraguay'),
    (76, 'Países Bajos', 'Marruecos'),
    (77, 'Costa de Marfil', 'Noruega'),
    (78, 'Francia', 'Suecia'),
    (79, 'México', 'Ecuador'),
    (80, 'Inglaterra', 'RD Congo'),
]
MATCHES_J8 = [
    (81, 'Bélgica', 'Senegal'),
    (82, 'Estados Unidos', 'Bosnia y Herzegovina'),
    (83, 'España', 'Austria'),
    (84, 'Portugal', 'Croacia'),
    (85, 'Suiza', 'Argelia'),
    (86, 'Australia', 'Egipto'),
    (87, 'Argentina', 'Cabo Verde'),
    (88, 'Colombia', 'Ghana'),
]

JORNADAS = (7, 8)
MATCHES = {7: MATCHES_J7, 8: MATCHES_J8}
# Hoja del Excel donde llegan las respuestas de cada ronda.
SHEET = {7: 'Form Responses 1', 8: 'j8'}
RONDA_TITULO = {
    7: 'Playoffs J7 · Dieciseisavos (primeros 8)',
    8: 'Playoffs J8 · Dieciseisavos (2ª tanda)',
}
RONDA_NOMBRE = {7: 'Cuartos', 8: 'Semifinales'}

# Form de predicciones de la ronda en curso (botón en la landing).
FORM_J8_URL = 'https://forms.gle/S7jBrrYNqR6XrNhs6'
CTA_FORM = (FORM_J8_URL, '📝 Llenar predicciones — Playoffs J8')

_CSS = """
  :root{--bg:#0b1020;--card:#141c2e;--bg2:#0f1523;--border:rgba(255,255,255,.08);
    --cyan:#00d4ff;--verde:#2ed573;--rojo:#ff4757;--dorado:#ffd700;--txt:#e6ecf7;
    --txt2:#8d99af;--gris:#5a6a80;}
  *{box-sizing:border-box;margin:0;padding:0;}
  body{font-family:-apple-system,"Segoe UI",Arial,sans-serif;background:var(--bg);color:var(--txt);}
  .hdr{background:linear-gradient(108deg,#00d8ec 0%,#1248c8 38%,#7218a8 66%,#3a0068 100%);
       padding:16px 22px;color:#fff;}
  .hdr-eyebrow{font-size:.74rem;font-weight:800;letter-spacing:1px;text-transform:uppercase;opacity:.9;}
  .hdr-llave{display:inline-block;margin-top:6px;font-size:.72rem;font-weight:800;letter-spacing:.5px;
       padding:3px 10px;border-radius:999px;background:rgba(255,255,255,.16);}
  .wrap{max-width:780px;margin:0 auto;padding:14px 12px 26px;}
  /* La grilla del header usa los MISMOS anchos que las columnas de la tabla
     (ver <colgroup>): así cada caricatura queda justo sobre su columna. */
  .vs{display:grid;grid-template-columns:34% 22% 22% 22%;align-items:center;margin:12px 0 6px;}
  .side{display:flex;flex-direction:column;align-items:center;gap:6px;min-width:0;}
  .av{width:78px;height:78px;border-radius:50%;object-fit:cover;border:3px solid var(--cyan);
      background:var(--card);}
  .side.b .av{border-color:var(--dorado);}
  .av-ph{width:78px;height:78px;border-radius:50%;display:flex;align-items:center;justify-content:center;
      font-weight:900;font-size:1.4rem;background:var(--card);border:3px solid var(--cyan);color:var(--txt2);}
  .side.b .av-ph{border-color:var(--dorado);}
  .nm{font-size:1.15rem;font-weight:900;color:var(--cyan);text-align:center;}
  .side.b .nm{color:var(--dorado);}
  .vs-mid{display:flex;flex-direction:column;align-items:center;gap:2px;}
  .vs-x{font-size:.8rem;font-weight:900;color:var(--txt2);}
  .vs-score{font-size:1.6rem;font-weight:900;color:var(--txt);letter-spacing:1px;white-space:nowrap;}
  .vs-score .a{color:var(--cyan);} .vs-score .b{color:var(--dorado);}
  .gana{font-size:.62rem;font-weight:800;letter-spacing:.5px;text-transform:uppercase;
        padding:2px 8px;border-radius:999px;background:rgba(46,213,115,.18);color:var(--verde);}
  table{width:100%;border-collapse:collapse;font-size:.8rem;margin-top:6px;}
  th,td{padding:6px 5px;border-bottom:1px solid var(--border);text-align:center;vertical-align:middle;}
  th{color:var(--txt2);font-size:.62rem;text-transform:uppercase;letter-spacing:.4px;font-weight:700;}
  th.real,td.real{background:rgba(255,255,255,.035);}
  th.pa,td.pa{text-align:left;color:var(--txt2);font-size:.74rem;}
  td.pa b{color:var(--txt);}
  .mk{font-size:1.02rem;font-weight:900;display:inline-block;}
  .mk.real{color:var(--txt);}
  .pg{display:block;font-size:.62rem;color:var(--txt2);margin-top:1px;}
  .porjugar{color:var(--gris);font-style:italic;font-size:.72rem;}
  .pend{color:var(--gris);font-style:italic;font-weight:700;font-size:.74rem;}
  .badge{display:inline-block;margin-top:5px;font-size:.78rem;font-weight:900;padding:2px 11px;
         border-radius:999px;border:1.5px solid transparent;line-height:1.25;}
  .badge.a{background:rgba(0,212,255,.22);color:var(--cyan);border-color:rgba(0,212,255,.65);
           box-shadow:0 0 11px rgba(0,212,255,.3);}
  .badge.b{background:rgba(255,215,0,.22);color:var(--dorado);border-color:rgba(255,215,0,.65);
           box-shadow:0 0 11px rgba(255,215,0,.3);}
  .badge.zero{background:rgba(141,153,175,.12);color:var(--gris);border-color:transparent;box-shadow:none;}
  tr.bono td{background:rgba(255,255,255,.03);}
  tr.bono .pa b{color:var(--cyan);}
  tr.tot td{border-top:2px solid var(--border);border-bottom:none;font-weight:900;}
  tr.tot .a{color:var(--cyan);font-size:1.05rem;} tr.tot .b{color:var(--dorado);font-size:1.05rem;}
  .nota{margin-top:14px;padding:11px 12px;border-radius:10px;font-size:.76rem;line-height:1.5;
    background:linear-gradient(180deg,rgba(0,212,255,.08),rgba(0,212,255,.02));
    border:1px solid rgba(0,212,255,.25);color:var(--txt2);}
  .nota b{color:var(--txt);}
  .foot{text-align:center;color:var(--gris);font-size:.72rem;padding:12px;}
  @media print{@page{size:A4;margin:0;}*{-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    .wrap{max-width:100%;}}
"""


# ── Datos ────────────────────────────────────────────────────────────────────
def _to_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _primer_gol_lv(equipo, local, visit):
    """Convierte el equipo del primer gol (nombre) a 'L'/'V'."""
    if equipo == local:
        return 'L'
    if equipo == visit:
        return 'V'
    return None


def _es_fila_resultado(nombre):
    """True si la fila es la de RESULTADOS reales (no un participante).
    Se captura escribiendo 'Respuesta'/'Resultado'/'Real' en la columna 'Tu nombre'."""
    return nombre is not None and str(nombre).strip().lower().startswith(
        ('respuesta', 'resultado', 'real'))


def _abrir_hoja(jornada):
    """Devuelve la hoja del Excel de la ronda, o None si no existe."""
    if not os.path.exists(PLAYOFF_XLSX):
        return None
    wb = openpyxl.load_workbook(PLAYOFF_XLSX, read_only=True, data_only=True)
    nombre = SHEET[jornada]
    # tolerante a mayúsculas ('j8' vs 'J8')
    real = next((s for s in wb.sheetnames if s.lower() == nombre.lower()), None)
    return wb[real] if real else None


def cargar_predicciones_playoff(jornada=7):
    """{nombre: {'marcadores': {num:(gl,gv,equipo_primer)}, 'rojas': int, 'penales': int}}."""
    ws = _abrir_hoja(jornada)
    preds = {}
    if ws is None:
        return preds
    from data.respuestas_loader import normalizar_nombre
    matches = MATCHES[jornada]
    for r in list(ws.iter_rows(values_only=True))[1:]:
        nombre = r[1] if len(r) > 1 else None
        if not nombre or _es_fila_resultado(nombre):
            continue  # fila vacía o la de resultados reales
        nombre = normalizar_nombre(nombre)  # 'Lu'→'Lucía', 'lucia'→'Lucía', etc.
        marc = {}
        for k, (num, _l, _v) in enumerate(matches):
            marc[num] = (_to_int(r[2 + 3 * k]), _to_int(r[3 + 3 * k]), r[4 + 3 * k])
        preds[nombre] = {
            'marcadores': marc,
            'rojas': _to_int(r[2 + 3 * len(matches)]),
            'penales': _to_int(r[3 + 3 * len(matches)]),
        }
    return preds


def _resultados_desde_excel(jornada=7):
    """Lee la fila de RESULTADOS reales de la hoja de la ronda (fila 'Respuesta')."""
    ws = _abrir_hoja(jornada)
    reales, rojas, penales = {}, None, None
    if ws is None:
        return reales, rojas, penales
    matches = MATCHES[jornada]
    for r in ws.iter_rows(values_only=True):
        if len(r) < 2 or not _es_fila_resultado(r[1]):
            continue
        for k, (num, local, visit) in enumerate(matches):
            gl, gv, eq = _to_int(r[2 + 3 * k]), _to_int(r[3 + 3 * k]), r[4 + 3 * k]
            if gl is None or gv is None:
                continue  # ese partido aún no tiene resultado
            pg = _primer_gol_lv(eq, local, visit) if (gl + gv) > 0 else None  # 0-0 → sin primer gol
            reales[num] = ResultadoPlayoff(num, gl, gv, pg)
        rojas = _to_int(r[2 + 3 * len(matches)])
        penales = _to_int(r[3 + 3 * len(matches)])
    return reales, rojas, penales


def cargar_resultados_playoff(jornada=7):
    """Devuelve ({num: ResultadoPlayoff}, total_rojas, total_penales).

    Fuente principal: la fila 'Respuesta' de la hoja de la ronda. Opcionalmente,
    data/resultados_playoffs.json puede sobreescribir partidos/totales puntuales.
    """
    reales, rojas, penales = _resultados_desde_excel(jornada)

    if os.path.exists(RESULTADOS_JSON):
        with open(RESULTADOS_JSON, encoding='utf-8') as f:
            data = json.load(f)
        j = data.get(str(jornada), {}) or {}
        for num, v in (j.get('marcadores') or {}).items():
            if not v:
                continue
            gl, gv, pg = (list(v) + [None, None, None])[:3]
            reales[int(num)] = ResultadoPlayoff(int(num), gl, gv, pg)
        if j.get('total_rojas') is not None:
            rojas = j['total_rojas']
        if j.get('total_penales') is not None:
            penales = j['total_penales']

    return reales, rojas, penales


def sincronizar_resultados_json():
    """Vuelca a data/resultados_playoffs.json los resultados que HOY tiene el Excel
    (fila 'Respuesta' de cada ronda). Se corre en cada `actualiza7` para dejar un
    respaldo versionado en git: cualquier cambio de resultado sale en el diff."""
    data = {}
    if os.path.exists(RESULTADOS_JSON):
        try:
            with open(RESULTADOS_JSON, encoding='utf-8') as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            data = {}
    data['_comment'] = ('Snapshot AUTO de los resultados del Excel de playoffs (fila '
                        "'Respuesta' de cada hoja). Se regenera en cada 'py actualiza7.py' "
                        'desde el Excel; sirve de respaldo versionado. Editarlo a mano no '
                        'sirve: el próximo run lo pisa con lo que diga el Excel.')
    for j in JORNADAS:
        if _abrir_hoja(j) is None:
            continue  # sin hoja para esa ronda → no tocar lo que ya haya
        reales, rojas, penales = _resultados_desde_excel(j)
        data[str(j)] = {
            'marcadores': {str(n): [r.goles_local, r.goles_visitante, r.primer_gol]
                           for n, r in sorted(reales.items())},
            'total_rojas': rojas,
            'total_penales': penales,
        }
    with open(RESULTADOS_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return data


def _seeds():
    tabla = calcular_tabla_general(cargar_participantes(), cargar_historial_resultados())
    return playoffs.sembrar(tabla)


def cruces(jornada=7):
    """[(jugadorA, jugadorB, llave)] de la ronda.

    J7 = cuartos (siembra directa). J8 = semifinales, resueltas con los puntos
    de la J7 (Campeones: avanza el ganador; Sótano: avanza el perdedor)."""
    seeds = _seeds()
    if jornada == 7:
        cr = [(a, b, 'campeones') for _id, a, b in playoffs._campeones_cuartos(seeds)]
        cr += [(a, b, 'sotano') for _id, a, b in playoffs._sotano_cuartos(seeds)]
        return cr
    if jornada == 8:
        pts7, _ = puntos_ronda(7)
        gan, perd = {}, {}
        for cid, a, b in playoffs._campeones_cuartos(seeds) + playoffs._sotano_cuartos(seeds):
            g, p = playoffs._h2h(a, b, pts7, seeds)
            gan[cid], perd[cid] = g, p
        cr = [(a, b, 'campeones') for _id, a, b in playoffs._campeones_semis(seeds, gan)]
        cr += [(a, b, 'sotano') for _id, a, b in playoffs._sotano_semis(seeds, perd)]
        return cr
    return []


def puntos_ronda(jornada=7):
    """Devuelve ({nombre: puntos de playoff en la ronda}, n_partidos_con_resultado)."""
    preds = cargar_predicciones_playoff(jornada)
    reales, rj, rp = cargar_resultados_playoff(jornada)
    matches = MATCHES[jornada]
    puntos = {}
    for n, d in preds.items():
        t = 0
        for num, local, visit in matches:
            real = reales.get(num)
            gl, gv, eq = d['marcadores'].get(num, (None, None, None))
            if real is None or gl is None:
                continue
            t += puntos_partido(
                PrediccionPlayoff('', num, gl, gv, _primer_gol_lv(eq, local, visit)), real)
        if rj is not None and d['rojas'] == rj:
            t += BONUS_ROJAS
        if rp is not None and d['penales'] == rp:
            t += BONUS_PENALES
        puntos[n] = t
    return puntos, len(reales)


def puntos_ronda_j7():
    """Compat: puntos de la J7 (usado por la proyección de playoffs)."""
    return puntos_ronda(7)


def _avatar_data_url(nombre):
    fname = _CARICATURA_FILE.get(nombre)
    if not fname:
        return None
    fpath = os.path.join(CARICATURAS_DIR, fname)
    if not os.path.exists(fpath):
        return None
    with open(fpath, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode()
    mime = 'image/png' if fname.lower().endswith('.png') else 'image/jpeg'
    return f'data:{mime};base64,{b64}'


# ── Render ───────────────────────────────────────────────────────────────────
def _avatar_html(nombre, lado):
    url = _avatar_data_url(nombre)
    if url:
        return f'<img class="av" src="{url}" alt="{nombre}">'
    ini = ''.join(w[0] for w in nombre.split()[:2]).upper()
    return f'<div class="av-ph">{ini}</div>'


def _celda_pred(pred, num, local, visit, real, lado):
    """(html, puntos|None) de la predicción de un jugador en un partido."""
    if pred is None:
        return '<span class="pend">— pendiente —</span>', None
    gl, gv, eq = pred['marcadores'].get(num, (None, None, None))
    if gl is None or gv is None:
        return '<span class="pend">—</span>', None
    html = f'<span class="mk">{gl}–{gv}</span><span class="pg">1er: {eq or "—"}</span>'
    pts = None
    if real is not None:
        po = PrediccionPlayoff('', num, gl, gv, _primer_gol_lv(eq, local, visit))
        pts = puntos_partido(po, real)
        cls = 'zero' if pts == 0 else lado
        html += f'<span class="badge {cls}">+{pts}</span>'
    return html, pts


def _celda_real(real, local, visit):
    if real is None:
        return '<span class="porjugar">por jugar</span>'
    eq = {'L': local, 'V': visit}.get(real.primer_gol, '— (0-0)')
    return (f'<span class="mk real">{real.goles_local}–{real.goles_visitante}</span>'
            f'<span class="pg">1er: {eq}</span>')


def _bono_pred(pred, clave, real_total, lado):
    if pred is None:
        return '<span class="pend">—</span>', None
    val = pred.get(clave)
    html = f'<span class="mk">{val if val is not None else "—"}</span>'
    pts = None
    if real_total is not None and val is not None:
        bonus = BONUS_ROJAS if clave == 'rojas' else BONUS_PENALES
        pts = bonus if val == real_total else 0
        cls = 'zero' if pts == 0 else lado
        html += f'<span class="badge {cls}">+{pts}</span>'
    return html, pts


def _seccion_cruce(a, b, llave, preds, reales, rojas_real, penales_real, jornada=7):
    matches = MATCHES[jornada]
    ronda_nom = RONDA_NOMBRE[jornada]
    pa, pb = preds.get(a), preds.get(b)
    llave_txt = (f'🏆 Campeones · {ronda_nom}' if llave == 'campeones'
                 else f'🚽 Sótano (Toilet Playoffs) · {ronda_nom}')
    tot_a = tot_b = 0
    hay_pts = False

    filas = ''
    for num, local, visit in matches:
        real = reales.get(num)
        ca, pa_pts = _celda_pred(pa, num, local, visit, real, 'a')
        cb, pb_pts = _celda_pred(pb, num, local, visit, real, 'b')
        if pa_pts is not None:
            tot_a += pa_pts; hay_pts = True
        if pb_pts is not None:
            tot_b += pb_pts; hay_pts = True
        filas += (
            f'<tr><td class="pa">#{num} <b>{local}</b> vs <b>{visit}</b></td>'
            f'<td>{ca}</td><td class="real">{_celda_real(real, local, visit)}</td><td>{cb}</td></tr>'
        )

    # bonos
    for clave, etq, real_total in (('rojas', '🟥 Total rojas', rojas_real),
                                   ('penales', '🎯 Total penales', penales_real)):
        ba, ba_pts = _bono_pred(pa, clave, real_total, 'a')
        bb, bb_pts = _bono_pred(pb, clave, real_total, 'b')
        if ba_pts is not None:
            tot_a += ba_pts; hay_pts = True
        if bb_pts is not None:
            tot_b += bb_pts; hay_pts = True
        real_txt = (f'<span class="mk real">{real_total}</span>'
                    if real_total is not None else '<span class="porjugar">—</span>')
        filas += (f'<tr class="bono"><td class="pa"><b>{etq}</b></td>'
                  f'<td>{ba}</td><td class="real">{real_txt}</td><td>{bb}</td></tr>')

    # total
    filas += (f'<tr class="tot"><td class="pa"><b>TOTAL</b></td>'
              f'<td class="a">{tot_a}</td><td class="real"></td><td class="b">{tot_b}</td></tr>')

    # marcador del cruce + líder
    score = f'<span class="a">{tot_a}</span> – <span class="b">{tot_b}</span>'
    lider = ''
    if hay_pts and tot_a != tot_b:
        ganador = a if tot_a > tot_b else b
        lider = f'<span class="gana">▲ {ganador} va arriba</span>'

    faltan = [n for n, p in ((a, pa), (b, pb)) if p is None]
    aviso = ''
    if faltan:
        aviso = (f'<div class="nota" style="border-color:rgba(255,71,87,.4);'
                 f'background:rgba(255,71,87,.07)">⏳ Falta(n) por enviar sus predicciones: '
                 f'<b>{", ".join(faltan)}</b>.</div>')

    html_doc = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{a} vs {b} — Playoffs J{jornada}</title><style>{_CSS}</style></head>
<body>
<header class="hdr">
  <div class="hdr-eyebrow">⚽ {RONDA_TITULO[jornada]}</div>
  <div class="hdr-llave">{llave_txt}</div>
</header>
<div class="wrap">
  <div class="vs">
    <div></div>
    <div class="side a">{_avatar_html(a, 'a')}<div class="nm">{a}</div></div>
    <div class="vs-mid"><span class="vs-x">VS</span><span class="vs-score">{score}</span>{lider}</div>
    <div class="side b">{_avatar_html(b, 'b')}<div class="nm">{b}</div></div>
  </div>
  <table>
    <colgroup><col style="width:34%"><col style="width:22%"><col style="width:22%"><col style="width:22%"></colgroup>
    <thead><tr><th class="pa">Partido</th><th>{a}</th><th class="real">Resultado real</th><th>{b}</th></tr></thead>
    <tbody>{filas}</tbody>
  </table>
  {aviso}
  <div class="nota">Puntos por partido: <b>marcador exacto = 3</b>, <b>solo resultado = 2</b>,
  <b>primer gol = +1</b> (máx 4). Bonos: <b>+2</b> rojas exactas, <b>+2</b> penales exactos.
  Avanza quien sume más puntos en la ronda. Los partidos "por jugar" aún no otorgan puntos.</div>
</div>
<div class="foot">Quiniela Mundial 2026 — Playoffs · cruce {a} vs {b}</div>
</body></html>"""
    return html_doc, tot_a, tot_b


def _slug(nombre):
    s = (nombre.lower().replace('á', 'a').replace('é', 'e').replace('í', 'i')
         .replace('ó', 'o').replace('ú', 'u'))
    return re.sub(r'[^a-z0-9]+', '_', s).strip('_')


_LANDING_CSS = """
  :root{--bg:#0b1020;--card:#141c2e;--bg2:#0f1523;--border:rgba(255,255,255,.08);
    --cyan:#00d4ff;--dorado:#ffd700;--txt:#e6ecf7;--txt2:#8d99af;--gris:#5a6a80;}
  *{box-sizing:border-box;margin:0;padding:0;}
  body{font-family:-apple-system,"Segoe UI",Arial,sans-serif;background:var(--bg);color:var(--txt);}
  .hdr{background:linear-gradient(108deg,#00d8ec 0%,#1248c8 38%,#7218a8 66%,#3a0068 100%);padding:18px 22px;color:#fff;}
  .hdr-title{font-size:1.5rem;font-weight:900;}
  .hdr-sub{font-size:.82rem;opacity:.9;margin-top:3px;}
  .wrap{max-width:880px;margin:0 auto;padding:14px 14px 30px;}
  .back{display:inline-block;margin:4px 0 10px;color:var(--cyan);text-decoration:none;font-weight:700;font-size:.82rem;}
  .navr{display:inline-block;margin:4px 8px 10px 0;color:var(--dorado);text-decoration:none;font-weight:800;font-size:.82rem;
        border:1px solid rgba(255,215,0,.4);border-radius:999px;padding:4px 12px;}
  .cta{display:block;margin:2px 0 6px;text-align:center;text-decoration:none;font-weight:900;font-size:.98rem;
       color:#001018;background:linear-gradient(90deg,#00d4ff,#2ed573);padding:13px 12px;border-radius:12px;}
  .cta:hover{filter:brightness(1.06);}
  .grupo-t{font-size:.78rem;font-weight:800;letter-spacing:.5px;text-transform:uppercase;color:var(--txt2);margin:16px 2px 9px;}
  .grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(250px,1fr));gap:12px;}
  .lc{display:block;background:var(--card);border:1px solid var(--border);border-radius:14px;
      padding:14px 12px;text-decoration:none;color:var(--txt);transition:border-color .15s;}
  .lc:hover{border-color:var(--cyan);}
  .lc.pend{opacity:.9;}
  .lc-vs{display:flex;align-items:center;justify-content:space-around;gap:6px;}
  .lc-side{display:flex;flex-direction:column;align-items:center;gap:5px;flex:1;min-width:0;}
  .lc .av{width:60px;height:60px;border-radius:50%;object-fit:cover;border:3px solid var(--cyan);background:var(--bg2);}
  .lc-side.b .av{border-color:var(--dorado);}
  .lc .av-ph{width:60px;height:60px;border-radius:50%;display:flex;align-items:center;justify-content:center;
         font-weight:900;background:var(--bg2);border:3px solid var(--cyan);color:var(--txt2);}
  .lc-side.b .av-ph{border-color:var(--dorado);}
  .lc-nm{font-size:.95rem;font-weight:900;color:var(--cyan);text-align:center;}
  .lc-nm.b{color:var(--dorado);}
  .lc-mid{display:flex;flex-direction:column;align-items:center;gap:1px;}
  .lc-x{font-size:.66rem;font-weight:900;color:var(--gris);letter-spacing:.5px;}
  .lc-score{font-size:1.5rem;font-weight:900;color:var(--txt);white-space:nowrap;line-height:1;}
  .lc-score .a{color:var(--cyan);} .lc-score .b{color:var(--dorado);}
  .lc-sclbl{font-size:.56rem;font-weight:700;letter-spacing:.5px;text-transform:uppercase;color:var(--gris);}
  .lc-pend{margin-top:10px;text-align:center;font-size:.72rem;font-weight:800;color:#ff6b81;
           background:rgba(255,71,87,.1);border-radius:999px;padding:3px 8px;}
  .lc-cta{margin-top:10px;text-align:center;font-size:.75rem;font-weight:800;color:var(--cyan);}
  .lc.pend .lc-cta{color:var(--txt2);}
  .foot{text-align:center;color:var(--gris);font-size:.72rem;padding:14px;}
"""


def _landing_html(info, jornada=7):
    """Página de acceso a los cruces (info = [(ruta, a, b, llave, faltan, ta, tb)])."""
    ronda_nom = RONDA_NOMBRE[jornada]
    grupos = {}
    for ruta, a, b, llave, faltan, ta, tb in info:
        grupos.setdefault(llave, []).append((os.path.basename(ruta), a, b, faltan, ta, tb))

    def _card(fname, a, b, faltan, ta, tb):
        pend = f'<div class="lc-pend">⏳ Falta: {", ".join(faltan)}</div>' if faltan else ''
        cta = 'Ver (incompleto) →' if faltan else 'Ver predicciones →'
        return f"""
        <a class="lc{' pend' if faltan else ''}" href="{fname}">
          <div class="lc-vs">
            <div class="lc-side">{_avatar_html(a, 'a')}<span class="lc-nm">{a}</span></div>
            <div class="lc-mid"><span class="lc-x">VS</span>
              <span class="lc-score"><b class="a">{ta}</b> – <b class="b">{tb}</b></span>
              <span class="lc-sclbl">puntos</span></div>
            <div class="lc-side b">{_avatar_html(b, 'b')}<span class="lc-nm b">{b}</span></div>
          </div>
          {pend}
          <div class="lc-cta">{cta}</div>
        </a>"""

    secciones = ''
    for llave, etq in (('campeones', f'🏆 Campeones · {ronda_nom}'),
                       ('sotano', f'🚽 Sótano (Toilet Playoffs) · {ronda_nom}')):
        if not grupos.get(llave):
            continue
        cards = ''.join(_card(*c) for c in grupos[llave])
        secciones += f'<div class="grupo-t">{etq}</div><div class="grid">{cards}</div>'

    # navegación entre rondas
    nav = ''
    for j in JORNADAS:
        if j != jornada:
            nav += f'<a class="navr" href="cruces_j{j}.html">Ver cruces J{j} →</a>'
    form_url, form_lbl = CTA_FORM

    return f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta http-equiv="Cache-Control" content="no-cache, must-revalidate">
<title>Cruces Playoffs J{jornada} — Quiniela Mundial 2026</title><style>{_LANDING_CSS}</style></head>
<body>
<header class="hdr">
  <div class="hdr-title">⚔️ Cruces Playoffs · Jornada {jornada}</div>
  <div class="hdr-sub">{ronda_nom} — predicciones head-to-head de cada cruce</div>
</header>
<div class="wrap">
  <a class="back" href="index.html">← Volver a la tabla</a>
  {nav}
  <a class="cta" href="{form_url}" target="_blank">{form_lbl}</a>
  {secciones}
</div>
<div class="foot">Quiniela Mundial 2026 — Playoffs J{jornada}</div>
</body></html>"""


def _escribir_cruces(output_dir, jornada=7):
    """Escribe un HTML por cruce en output_dir. [(ruta, a, b, llave, faltan, ta, tb)]."""
    preds = cargar_predicciones_playoff(jornada)
    reales, rojas_real, penales_real = cargar_resultados_playoff(jornada)
    os.makedirs(output_dir, exist_ok=True)
    info = []
    for a, b, llave in cruces(jornada):
        html, ta, tb = _seccion_cruce(a, b, llave, preds, reales, rojas_real, penales_real, jornada)
        fname = f'h2h_j{jornada}_{_slug(a)}_vs_{_slug(b)}.html'
        with open(os.path.join(output_dir, fname), 'w', encoding='utf-8') as f:
            f.write(html)
        faltan = [x for x in (a, b) if preds.get(x) is None]
        info.append((os.path.join(output_dir, fname), a, b, llave, faltan, ta, tb))
    return info


def generar(output_dir=OUTPUT_DIR, jornada=7):
    """Genera los cruces de la ronda. [(ruta,a,b,llave,completos)]."""
    return [(r, a, b, llave, not faltan)
            for r, a, b, llave, faltan, ta, tb in _escribir_cruces(output_dir, jornada)]


def generar_web(jornada=7):
    """Genera los cruces de la ronda en docs/ + la página docs/cruces_j{n}.html."""
    info = _escribir_cruces(DOCS_DIR, jornada)
    landing = _landing_html(info, jornada)
    lpath = os.path.join(DOCS_DIR, f'cruces_j{jornada}.html')
    with open(lpath, 'w', encoding='utf-8') as f:
        f.write(landing)
    return lpath, info


if __name__ == '__main__':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except AttributeError:
        pass
    primera = None
    for j in JORNADAS:
        rutas = generar(jornada=j)
        print(f'J{j}: {len(rutas)} cruce(s) en {OUTPUT_DIR}')
        for ruta, a, b, llave, completos in rutas:
            estado = 'completo' if completos else 'FALTA alguien'
            print(f'   [{llave:9}] {a} vs {b:12} -> {os.path.basename(ruta)}  ({estado})')
            primera = primera or ruta
    if primera:
        webbrowser.open(f'file:///{primera.replace(os.sep, "/")}')
